"""
Stage 5: clean rows -> Master Excel file.

Each pipeline run (one brand) appends its rows to the same master workbook,
replacing any existing rows for that brand+version so re-running a brand
doesn't create duplicates. Formatting (column widths, header styling, frozen
header row) is applied once per write so the file stays readable as it grows
to dozens of brands.

The write is atomic: everything goes to a temp file first, which then
replaces the master file in one step. A crash mid-write can no longer
corrupt the accumulated master data.
"""
from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import MASTER_COLUMNS, MASTER_EXCEL_PATH

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _load_existing(path: Path) -> pd.DataFrame:
    if path.exists():
        return pd.read_excel(path, sheet_name="Master", dtype=object)
    return pd.DataFrame(columns=MASTER_COLUMNS)


def _style_workbook(path: Path) -> None:
    """Apply header styling, frozen header row, and sensible column widths."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["Master"]

    for col_idx, column_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

        # Width heuristic: header length vs a reasonable cap, so "Notes"
        # doesn't end up either truncated or absurdly wide.
        width = min(max(len(column_name) + 2, 12), 32)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)


def _as_text(series: pd.Series) -> pd.Series:
    """
    Normalize a column for identity comparison. Empty cells round-trip
    through Excel as NaN, so NaN and "" must compare as equal — otherwise
    re-running a brand with the default (empty) version label would fail to
    replace its old rows and silently duplicate them.
    """
    return series.fillna("").astype(str)


def write_master_excel(
    new_rows: list[dict],
    brand: str,
    price_list_version: str,
    path: Path = MASTER_EXCEL_PATH,
) -> Path:
    """
    Merge new_rows into the master Excel file.

    Any existing rows for the same (Brand, Price List Version) are replaced —
    this is what makes re-running a brand's pipeline safe to do repeatedly
    rather than accumulating duplicate copies on every run.
    """
    existing = _load_existing(path)

    if not existing.empty:
        is_same_brand_version = (
            (_as_text(existing["Brand"]) == str(brand))
            & (_as_text(existing["Price List Version"]) == str(price_list_version))
        )
        existing = existing[~is_same_brand_version]

    new_df = pd.DataFrame(new_rows, columns=MASTER_COLUMNS)

    if existing.empty:
        combined = new_df
    else:
        combined = pd.concat(
            [existing.astype(object), new_df.astype(object)], ignore_index=True
        )
    combined = combined[MASTER_COLUMNS]  # enforce column order

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name="Master", index=False)
        _style_workbook(tmp_path)
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    return path
